"""
Smoke tests for the excel_audit_export module.

Builds a run archive (and optionally a comparison bundle) in a temp dir,
exports the audit workbook, then re-opens the ``.xlsx`` with openpyxl to
verify expected sheet names, headers, row counts, and that no value in
any table leaks out as an Excel formula error (``#REF``, ``#NAME``, etc.).
"""

from __future__ import annotations

import os
import sys
import tempfile

import pandas as pd

_HERE = os.path.dirname(os.path.abspath(__file__))
_STREAMLIT_DIR = os.path.abspath(os.path.join(_HERE, ".."))
if _STREAMLIT_DIR not in sys.path:
    sys.path.insert(0, _STREAMLIT_DIR)

from openpyxl import load_workbook  # noqa: E402

from run_archive import create_run_archive  # noqa: E402
from run_comparison import run_comparison  # noqa: E402
from excel_audit_export import (  # noqa: E402
    SHEET_README,
    SHEET_METHODOLOGY,
    SHEET_DATA_QUALITY,
    SHEET_SCORED,
    SHEET_TOP50_2023,
    SHEET_TOP50_2025,
    SHEET_TOP50_CONSENSUS,
    SHEET_DISAGREEMENT,
    SHEET_WHAT_CHANGED,
    SHEET_SCORE_MOVERS,
    SHEET_BAND_CHANGES,
    SHEET_QUADRANT_CHANGES,
    SHEET_ACTION_CHANGES,
    SHEET_NEW_FUNDS,
    SHEET_REMOVED_FUNDS,
    StaleScoreArchiveError,
    _default_workbook_path,
    build_audit_workbook,
    export_latest_run,
    export_run,
    validate_score_bounds,
)

EXCEL_ERRORS = {"#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#N/A", "#NULL!", "#NUM!"}


def _synthetic_table() -> pd.DataFrame:
    rows = []
    for i in range(6):
        rows.append({
            "Symbol": f"FUN{i}",
            "Name": f"Fund {i}",
            "Category": "Large Growth",
            "Fund_Type": "Passive" if i % 2 == 0 else "Active",
            "Score_2023_Final": 80.0 - i * 7.5,
            "Score_2025_Final": 82.0 - i * 6.0,
            "Score_Gap": 2.0 + i * 1.5,
            "Rank_2023": i + 1,
            "Rank_2025": i + 1,
            "Consensus_Rank": i + 1,
            "Score_Band_2023": "STRONG" if (80.0 - i * 7.5) >= 80 else ("REVIEW" if (80.0 - i * 7.5) >= 60 else "WEAK"),
            "Score_Band_2025": "STRONG" if (82.0 - i * 6.0) >= 80 else ("REVIEW" if (82.0 - i * 6.0) >= 60 else "WEAK"),
            "Quadrant": "Q1_Both_Strong" if (80.0 - i * 7.5) >= 80 and (82.0 - i * 6.0) >= 80 else "Q4_Both_Weak",
            "Data_Coverage_2023": 0.9,
            "Data_Coverage_2025": 0.95,
            "Primary_Driver": "Stable",
            "Action_Flag": "LEAD" if i == 0 else "WATCH",
        })
    return pd.DataFrame(rows)


def _assert_no_formula_errors(ws) -> None:
    for row in ws.iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and value in EXCEL_ERRORS:
                raise AssertionError(
                    f"Excel error value {value!r} found in sheet {ws.title!r}"
                )


def _headers(ws) -> list:
    return [cell.value for cell in ws[1]]


def test_build_workbook_has_expected_base_sheets():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_latest_run(runs_dir=tmp)
        assert out.endswith(".xlsx") and os.path.isfile(out)
        assert out == _default_workbook_path(os.path.join(tmp, "2026-04-10"))

        wb = load_workbook(out)
        expected = {
            SHEET_README, SHEET_METHODOLOGY, SHEET_DATA_QUALITY,
            SHEET_SCORED, SHEET_TOP50_2023, SHEET_TOP50_2025,
            SHEET_TOP50_CONSENSUS, SHEET_DISAGREEMENT,
        }
        assert expected.issubset(set(wb.sheetnames))
        # No What_Changed when only a single run is archived.
        assert SHEET_WHAT_CHANGED not in wb.sheetnames


def test_scored_funds_table_matches_source_row_count():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_run(run_date="2026-04-10", runs_dir=tmp)
        wb = load_workbook(out)
        ws = wb[SHEET_SCORED]

        header = _headers(ws)
        for col in ("Symbol", "Score_2023_Final", "Score_2025_Final",
                    "Score_Band_2023", "Score_Band_2025", "Quadrant",
                    "Action_Flag"):
            assert col in header, f"missing {col} in Scored_Funds header"

        # header row + one row per fund
        assert ws.max_row == len(table) + 1
        _assert_no_formula_errors(ws)


def test_top_n_sheets_are_bounded_and_sorted():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_latest_run(runs_dir=tmp)
        wb = load_workbook(out)

        for sheet_name, score_col in (
            (SHEET_TOP50_2023, "Score_2023_Final"),
            (SHEET_TOP50_2025, "Score_2025_Final"),
        ):
            ws = wb[sheet_name]
            header = _headers(ws)
            assert score_col in header
            # Top-N never exceeds the source fund count.
            assert ws.max_row - 1 <= len(table)
            idx = header.index(score_col) + 1
            values = [ws.cell(row=r, column=idx).value for r in range(2, ws.max_row + 1)]
            numeric = [v for v in values if isinstance(v, (int, float))]
            assert numeric == sorted(numeric, reverse=True)
            _assert_no_formula_errors(ws)

        ws_c = wb[SHEET_TOP50_CONSENSUS]
        assert "Consensus_Rank" in _headers(ws_c)
        _assert_no_formula_errors(ws_c)


def test_readme_contains_run_date_and_row_counts():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_latest_run(runs_dir=tmp)
        wb = load_workbook(out)
        ws = wb[SHEET_README]

        text = []
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    text.append(str(value))
        blob = "\n".join(text)
        assert "2026-04-10" in blob
        assert "Scored funds" in blob
        assert "Score Bands" in blob


def test_data_quality_sheet_has_core_stats():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_latest_run(runs_dir=tmp)
        wb = load_workbook(out)
        ws = wb[SHEET_DATA_QUALITY]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Row count" in labels
        assert "Joined count (both systems)" in labels
        assert "score_2023.mean" in labels
        _assert_no_formula_errors(ws)


def test_no_formula_errors_anywhere():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        out = export_latest_run(runs_dir=tmp)
        wb = load_workbook(out)
        for name in wb.sheetnames:
            _assert_no_formula_errors(wb[name])


def test_disagreement_sheet_includes_band_mismatches():
    df = pd.DataFrame([
        {
            "Symbol": "AGREE", "Name": "Agree Fund", "Category": "X",
            "Fund_Type": "Passive",
            "Score_2023_Final": 85.0, "Score_2025_Final": 86.0,
            "Score_Gap": 1.0,
            "Rank_2023": 1, "Rank_2025": 1, "Consensus_Rank": 1,
            "Score_Band_2023": "STRONG", "Score_Band_2025": "STRONG",
            "Quadrant": "Q1_Both_Strong",
            "Data_Coverage_2023": 1.0, "Data_Coverage_2025": 1.0,
            "Primary_Driver": "Stable", "Action_Flag": "LEAD",
        },
        {
            "Symbol": "DISAGREE", "Name": "Disagree Fund", "Category": "X",
            "Fund_Type": "Active",
            "Score_2023_Final": 45.0, "Score_2025_Final": 85.0,
            "Score_Gap": 40.0,
            "Rank_2023": 2, "Rank_2025": 2, "Consensus_Rank": 2,
            "Score_Band_2023": "WEAK", "Score_Band_2025": "STRONG",
            "Quadrant": "Q2_Only_2025",
            "Data_Coverage_2023": 0.8, "Data_Coverage_2025": 0.9,
            "Primary_Driver": "Upgraded by 2025 system",
            "Action_Flag": "REVIEW",
        },
    ])
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=df)
        out = export_latest_run(runs_dir=tmp)
        wb = load_workbook(out)
        ws = wb[SHEET_DISAGREEMENT]
        header = _headers(ws)
        sym_idx = header.index("Symbol") + 1
        symbols = [ws.cell(row=r, column=sym_idx).value for r in range(2, ws.max_row + 1)]
        assert "DISAGREE" in symbols
        assert "AGREE" not in symbols


def test_comparison_sheets_emitted_when_comparison_exists():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-03-10", runs_dir=tmp, table=table)
        # Shift a single score so the comparison isn't empty.
        shifted = table.copy()
        shifted.loc[0, "Score_2025_Final"] = 75.0
        shifted.loc[0, "Score_Band_2025"] = "REVIEW"
        shifted.loc[0, "Quadrant"] = "Q4_Both_Weak"
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=shifted)

        run_comparison(runs_dir=tmp, write=True)
        out = export_latest_run(runs_dir=tmp)

        wb = load_workbook(out)
        for name in (SHEET_WHAT_CHANGED, SHEET_SCORE_MOVERS, SHEET_BAND_CHANGES,
                     SHEET_QUADRANT_CHANGES, SHEET_ACTION_CHANGES,
                     SHEET_NEW_FUNDS, SHEET_REMOVED_FUNDS):
            assert name in wb.sheetnames, f"missing {name}"
            _assert_no_formula_errors(wb[name])


def test_no_comparison_flag_skips_comparison_sheets():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-03-10", runs_dir=tmp, table=table)
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        run_comparison(runs_dir=tmp, write=True)
        out = export_latest_run(runs_dir=tmp, include_comparison=False)
        wb = load_workbook(out)
        assert SHEET_WHAT_CHANGED not in wb.sheetnames
        assert SHEET_SCORE_MOVERS not in wb.sheetnames


def test_build_audit_workbook_accepts_explicit_out_path():
    table = _synthetic_table()
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        from run_archive import load_latest_run
        run = load_latest_run(tmp)
        custom = os.path.join(tmp, "nested", "custom.xlsx")
        target = build_audit_workbook(run=run, out_path=custom, runs_dir=tmp)
        assert target == custom
        assert os.path.isfile(custom)


def test_validate_score_bounds_passes_on_clean_table():
    table = _synthetic_table()
    validate_score_bounds(table)  # must not raise


def test_validate_score_bounds_flags_passive_over_100():
    table = _synthetic_table()
    table.loc[0, "Score_2025_Final"] = 109.0
    try:
        validate_score_bounds(table)
    except StaleScoreArchiveError as exc:
        msg = str(exc)
        assert "Score_2025_Final" in msg
        assert "109" in msg
        assert "overwrite" in msg.lower()
    else:
        raise AssertionError("Expected StaleScoreArchiveError, got none")


def test_validate_score_bounds_flags_2023_over_100():
    table = _synthetic_table()
    table.loc[2, "Score_2023_Final"] = 115.0
    try:
        validate_score_bounds(table)
    except StaleScoreArchiveError as exc:
        assert "Score_2023_Final" in str(exc)
    else:
        raise AssertionError("Expected StaleScoreArchiveError, got none")


def test_build_audit_workbook_refuses_stale_archive():
    """If an archived CSV has scores >100, the export must refuse rather than
    ship a workbook that violates the 0-100 invariant."""
    table = _synthetic_table()
    table.loc[0, "Score_2025_Final"] = 108.5
    with tempfile.TemporaryDirectory() as tmp:
        create_run_archive(run_date="2026-04-10", runs_dir=tmp, table=table)
        try:
            export_run(run_date="2026-04-10", runs_dir=tmp)
        except StaleScoreArchiveError as exc:
            assert "108" in str(exc) or "Score_2025_Final" in str(exc)
        else:
            raise AssertionError("Expected StaleScoreArchiveError, got none")


def main() -> int:
    funcs = [
        test_build_workbook_has_expected_base_sheets,
        test_scored_funds_table_matches_source_row_count,
        test_top_n_sheets_are_bounded_and_sorted,
        test_readme_contains_run_date_and_row_counts,
        test_data_quality_sheet_has_core_stats,
        test_no_formula_errors_anywhere,
        test_disagreement_sheet_includes_band_mismatches,
        test_comparison_sheets_emitted_when_comparison_exists,
        test_no_comparison_flag_skips_comparison_sheets,
        test_build_audit_workbook_accepts_explicit_out_path,
        test_validate_score_bounds_passes_on_clean_table,
        test_validate_score_bounds_flags_passive_over_100,
        test_validate_score_bounds_flags_2023_over_100,
        test_build_audit_workbook_refuses_stale_archive,
    ]
    failed = 0
    for fn in funcs:
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except AssertionError as e:
            failed += 1
            print(f"FAIL {fn.__name__}: {e}")
        except Exception as e:  # noqa: BLE001
            failed += 1
            print(f"ERROR {fn.__name__}: {type(e).__name__}: {e}")
    return failed


if __name__ == "__main__":
    sys.exit(main())
