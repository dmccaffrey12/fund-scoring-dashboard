"""Integration: overlay artifacts wired through workflow_ui + Excel audit."""

from __future__ import annotations

import os
import sys
import tempfile

import pandas as pd
import pytest

_HERE = os.path.dirname(os.path.abspath(__file__))
_STREAMLIT_DIR = os.path.abspath(os.path.join(_HERE, ".."))
_FIXTURES = os.path.join(_HERE, "fixtures")
if _STREAMLIT_DIR not in sys.path:
    sys.path.insert(0, _STREAMLIT_DIR)

from dual_score_table import build_dual_score_table  # noqa: E402
from excel_audit_export import build_audit_workbook  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from run_archive import create_run_archive, load_run, run_overlay_dir  # noqa: E402
import workflow_ui  # noqa: E402


def test_generate_and_load_overlay_via_workflow_ui(tmp_path):
    runs_dir = str(tmp_path / "runs")
    # Build a full archive from bundled sample data so we have a real
    # dual-score table with enough coverage to join against.
    table = build_dual_score_table()
    create_run_archive(
        run_date="2026-04-22",
        runs_dir=runs_dir,
        table=table,
    )

    # Holdings file — reuse the sanitized fixture for the overlay.
    holdings_path = os.path.join(_FIXTURES, "model_holdings_good.csv")

    bundle = workflow_ui.generate_model_overlay_for_run(
        holdings_path=holdings_path,
        run_date="2026-04-22",
        runs_dir=runs_dir,
    )
    assert bundle["metadata"]["model_count"] >= 1
    assert not bundle["scorecard"].empty

    # Loading from disk returns the same scorecard row count.
    loaded = workflow_ui.load_model_overlay_for_run(
        "2026-04-22", runs_dir=runs_dir,
    )
    assert loaded is not None
    assert len(loaded["scorecard"]) == len(bundle["scorecard"])

    # Persisted under runs/<date>/model_holdings/ exactly.
    expected_dir = run_overlay_dir(runs_dir, "2026-04-22")
    assert os.path.isdir(expected_dir)
    assert os.path.isfile(os.path.join(expected_dir, "model_holdings_scorecard.csv"))


def test_excel_audit_includes_overlay_sheets(tmp_path):
    runs_dir = str(tmp_path / "runs")
    table = build_dual_score_table()
    create_run_archive(
        run_date="2026-04-22", runs_dir=runs_dir, table=table,
    )
    holdings_path = os.path.join(_FIXTURES, "model_holdings_good.csv")
    workflow_ui.generate_model_overlay_for_run(
        holdings_path=holdings_path,
        run_date="2026-04-22",
        runs_dir=runs_dir,
    )

    run = load_run("2026-04-22", runs_dir=runs_dir)
    out_xlsx = str(tmp_path / "audit.xlsx")
    build_audit_workbook(run=run, out_path=out_xlsx, runs_dir=runs_dir)

    wb = load_workbook(out_xlsx, read_only=True)
    sheets = set(wb.sheetnames)
    assert "Model_Summary" in sheets
    assert "Model_Holdings" in sheets
    assert "Model_Review_List" in sheets
    assert "Research_Candidates" in sheets
    assert "Replacement_Candidates" in sheets


def test_overlay_includes_alias_columns_in_persisted_scorecard(tmp_path):
    runs_dir = str(tmp_path / "runs")
    table = build_dual_score_table()
    create_run_archive(
        run_date="2026-04-22", runs_dir=runs_dir, table=table,
    )
    holdings_path = os.path.join(_FIXTURES, "model_holdings_good.csv")
    workflow_ui.generate_model_overlay_for_run(
        holdings_path=holdings_path,
        run_date="2026-04-22",
        runs_dir=runs_dir,
    )
    # Persisted scorecard must carry the new reconciliation columns so
    # downstream readers (Excel, committee review) see both the original
    # committee-facing Symbol and the resolved Scoring_Symbol.
    scorecard_path = os.path.join(
        run_overlay_dir(runs_dir, "2026-04-22"),
        "model_holdings_scorecard.csv",
    )
    sc = pd.read_csv(scorecard_path)
    for col in ("Symbol", "Scoring_Symbol", "Alias_Applied"):
        assert col in sc.columns


def test_excel_audit_skips_overlay_sheets_when_absent(tmp_path):
    runs_dir = str(tmp_path / "runs")
    table = build_dual_score_table()
    create_run_archive(
        run_date="2026-04-22", runs_dir=runs_dir, table=table,
    )
    # No overlay created — workbook must still build.
    run = load_run("2026-04-22", runs_dir=runs_dir)
    out_xlsx = str(tmp_path / "audit.xlsx")
    build_audit_workbook(run=run, out_path=out_xlsx, runs_dir=runs_dir)
    wb = load_workbook(out_xlsx, read_only=True)
    assert "Model_Summary" not in wb.sheetnames
    assert "Model_Holdings" not in wb.sheetnames
