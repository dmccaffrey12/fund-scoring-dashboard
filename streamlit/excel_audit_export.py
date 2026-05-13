"""
Excel Audit Workbook Export
===========================
Generate an Excel 2019-compatible ``.xlsx`` audit workbook from an
archived scoring run (produced by ``run_archive.py``).

The target consumer is an Excel 2019 user on the committee side — that
means we deliberately avoid anything that requires Microsoft 365 /
Excel 2021 dynamic-array engine:

    * no ``XLOOKUP``, ``LET``, ``FILTER``, ``SORT``, ``UNIQUE``
    * no spilled-range references (``Table1[#Data]`` is fine; ``SEQUENCE``
      is not)
    * every value the analyst sees is precomputed — formulas are the
      exception, not the rule

The workbook is intentionally self-contained: it bundles the canonical
dual-score table, ranked Top-50 cuts, disagreement and quadrant views,
plus generation metadata, validation counters, and (when available) the
month-over-month comparison tables from ``run_comparison.py``.

Default output path::

    runs/YYYY-MM-DD/reports/fundscore_audit_workbook.xlsx

Public API
----------
``build_audit_workbook(...)``   — write the workbook to a path.
``export_latest_run(...)``       — convenience: latest archive → workbook.
``export_run(run_date, ...)``    — convenience: specific archive → workbook.

CLI
---
``python streamlit/excel_audit_export.py export``   (latest run by default)
``python streamlit/excel_audit_export.py export --run-date 2026-04-30``
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from run_archive import (
    DEFAULT_RUNS_DIR,
    list_runs,
    load_latest_run,
    load_run,
)

# ---------------------------------------------------------------------------
# Defaults & constants
# ---------------------------------------------------------------------------

REPORTS_SUBDIR = "reports"
DEFAULT_WORKBOOK_NAME = "fundscore_audit_workbook.xlsx"
TOP_N = 50

WORKBOOK_SCHEMA_VERSION = "1.0"

# Excel enforces 31-char sheet names and disallows some punctuation; keep
# our names short and plain ASCII so they survive a round-trip through
# legacy Windows Explorer / SharePoint.
SHEET_README = "README_Runbook"
SHEET_METHODOLOGY = "Methodology"
SHEET_DATA_QUALITY = "Data_Quality"
SHEET_SCORED = "Scored_Funds"
SHEET_TOP50_2023 = "Top_50_2023"
SHEET_TOP50_2025 = "Top_50_2025"
SHEET_TOP50_CONSENSUS = "Top_50_Consensus"
SHEET_DISAGREEMENT = "Disagreement_List"
SHEET_WHAT_CHANGED = "What_Changed"
SHEET_SCORE_MOVERS = "Score_Movers"
SHEET_BAND_CHANGES = "Band_Changes"
SHEET_QUADRANT_CHANGES = "Quadrant_Changes"
SHEET_ACTION_CHANGES = "Action_Flag_Changes"
SHEET_NEW_FUNDS = "New_Funds"
SHEET_REMOVED_FUNDS = "Removed_Funds"

# Model-holdings overlay sheets — added only when the overlay bundle is
# present on disk under <run>/model_holdings/.
SHEET_MODEL_SUMMARY = "Model_Summary"
SHEET_MODEL_HOLDINGS = "Model_Holdings"
SHEET_MODEL_REVIEW_LIST = "Model_Review_List"
SHEET_RESEARCH_CANDIDATES = "Research_Candidates"
SHEET_REPLACEMENT_CANDIDATES = "Replacement_Candidates"

COMPARISON_SUBDIR = "comparison"  # matches run_comparison.default_comparison_dir
MODEL_OVERLAY_SUBDIR = "model_holdings"  # matches model_holdings_overlay.OVERLAY_SUBDIR

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
SECTION_FONT = Font(name="Calibri", size=13, bold=True, color="FF1F4E78")
BODY_FONT = Font(name="Calibri", size=11)
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
THIN = Side(border_style="thin", color="FFB7B7B7")
BORDER_BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BAND_FILL = {
    "STRONG": PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid"),
    "REVIEW": PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid"),
    "WEAK": PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _default_workbook_path(run_dir: str) -> str:
    return os.path.join(run_dir, REPORTS_SUBDIR, DEFAULT_WORKBOOK_NAME)


def _sanitize_table_name(name: str) -> str:
    """Excel Table names cannot contain spaces or start with a digit."""
    safe = "".join(ch if (ch.isalnum() or ch == "_") else "_" for ch in name)
    if not safe or safe[0].isdigit():
        safe = "T_" + safe
    return safe[:255]


def _coerce_cell(value: Any) -> Any:
    """Make a value safe to pass straight to openpyxl.

    openpyxl accepts basic Python scalars, datetimes, and bools. It will
    reject numpy scalars via deprecation paths and will happily write
    NaN into an XLSX as the literal float NaN — Excel then shows ``#NUM``.
    Convert NaN/None to empty string; cast numpy scalars to their Python
    equivalents.
    """
    if value is None:
        return ""
    # pandas NA / NaN floats
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    # numpy scalars expose .item()
    item = getattr(value, "item", None)
    if callable(item):
        try:
            return item()
        except (ValueError, TypeError):
            return value
    return value


def _style_header_row(ws: Worksheet, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER_BOX


def _auto_size_columns(ws: Worksheet, df: pd.DataFrame, start_col: int = 1) -> None:
    for idx, col in enumerate(df.columns):
        header_len = len(str(col))
        try:
            sample = df[col].astype(str).head(200)
            body_len = int(sample.map(len).max()) if len(sample) else 0
        except Exception:  # noqa: BLE001
            body_len = 0
        width = min(max(header_len, body_len) + 2, 48)
        ws.column_dimensions[get_column_letter(start_col + idx)].width = max(width, 10)


def _add_table(
    ws: Worksheet,
    df: pd.DataFrame,
    table_name: str,
    start_row: int = 1,
) -> None:
    """Write a DataFrame as a formatted Excel Table (Excel 2019 safe).

    Uses Excel's built-in Table feature (``xl/tables/*.xml``), which is
    supported all the way back to Excel 2007. Table styles do not rely on
    dynamic-array formulas, so legacy Excel renders them correctly.
    """
    if df is None or df.empty:
        # Excel refuses to create a zero-row table; write a single
        # placeholder row so the sheet is still navigable.
        df = pd.DataFrame({col: ["(no rows)"] for col in (df.columns if df is not None else ["Note"])})

    n_rows, n_cols = df.shape

    for c_idx, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=c_idx, value=str(col))
    _style_header_row(ws, start_row, n_cols)

    for r_idx, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for c_idx, col in enumerate(df.columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_coerce_cell(row[col]))

    last_col_letter = get_column_letter(n_cols)
    ref = f"A{start_row}:{last_col_letter}{start_row + n_rows}"
    table = Table(displayName=_sanitize_table_name(table_name), ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    _auto_size_columns(ws, df)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def _apply_band_fills(
    ws: Worksheet,
    df: pd.DataFrame,
    band_columns: Tuple[str, ...],
    start_row: int = 2,
) -> None:
    """Shade STRONG/REVIEW/WEAK cells so the committee can scan quickly."""
    for col_name in band_columns:
        if col_name not in df.columns:
            continue
        col_idx = list(df.columns).index(col_name) + 1
        for r_offset, value in enumerate(df[col_name].tolist()):
            fill = BAND_FILL.get(str(value))
            if fill is not None:
                ws.cell(row=start_row + r_offset, column=col_idx).fill = fill


# ---------------------------------------------------------------------------
# Content builders
# ---------------------------------------------------------------------------

def _top_n_by_score(table: pd.DataFrame, score_col: str, n: int = TOP_N) -> pd.DataFrame:
    if score_col not in table.columns:
        return pd.DataFrame()
    sub = table.dropna(subset=[score_col]).copy()
    # Stable sort with Symbol tie-break for deterministic Top-N ordering.
    sort_cols = [score_col, "Symbol"] if "Symbol" in sub.columns else [score_col]
    ascending = [False, True] if "Symbol" in sub.columns else [False]
    sub = sub.sort_values(
        sort_cols, ascending=ascending, kind="stable",
    ).head(n).reset_index(drop=True)
    rank_col = f"Rank_By_{score_col}"
    sub.insert(0, rank_col, range(1, len(sub) + 1))
    return sub


def _top_n_by_consensus(table: pd.DataFrame, n: int = TOP_N) -> pd.DataFrame:
    if "Consensus_Rank" not in table.columns:
        return pd.DataFrame()
    sub = table.dropna(subset=["Consensus_Rank"]).copy()
    sort_cols = ["Consensus_Rank", "Symbol"] if "Symbol" in sub.columns else ["Consensus_Rank"]
    ascending = [True, True] if "Symbol" in sub.columns else [True]
    sub = sub.sort_values(
        sort_cols, ascending=ascending, kind="stable",
    ).head(n).reset_index(drop=True)
    return sub


def _disagreement_list(table: pd.DataFrame) -> pd.DataFrame:
    """Rows where the two systems disagree (|Score_Gap| >= 10 or band diff)."""
    if table.empty:
        return table
    df = table.copy()
    gap_mask = pd.Series(False, index=df.index)
    if "Score_Gap" in df.columns:
        gap_mask = df["Score_Gap"].abs().fillna(0) >= 10
    band_mask = pd.Series(False, index=df.index)
    if {"Score_Band_2023", "Score_Band_2025"}.issubset(df.columns):
        band_mask = df["Score_Band_2023"].fillna("") != df["Score_Band_2025"].fillna("")
    picked = df[gap_mask | band_mask].copy()
    if "Score_Gap" in picked.columns:
        picked["_abs_gap"] = picked["Score_Gap"].abs()
        sort_cols = ["_abs_gap", "Symbol"] if "Symbol" in picked.columns else ["_abs_gap"]
        ascending = [False, True] if "Symbol" in picked.columns else [False]
        picked = picked.sort_values(
            sort_cols, ascending=ascending, kind="stable",
        ).drop(columns="_abs_gap")
    return picked.reset_index(drop=True)


def _load_comparison_tables(
    runs_dir: str, run_date: str
) -> Optional[Dict[str, Any]]:
    """Return the newest comparison bundle under the given run, if any.

    ``run_comparison`` writes to ``<run>/comparison/prior_<prior>/``.
    We pick the directory with the lexicographically greatest ``prior_*``
    folder (i.e. the most recent prior available).
    """
    base = os.path.join(runs_dir, run_date, COMPARISON_SUBDIR)
    if not os.path.isdir(base):
        return None
    candidates = sorted(
        name for name in os.listdir(base)
        if name.startswith("prior_")
        and os.path.isdir(os.path.join(base, name))
    )
    if not candidates:
        return None
    chosen = candidates[-1]
    folder = os.path.join(base, chosen)
    prior_date = chosen[len("prior_"):]

    def _safe_read(fname: str) -> pd.DataFrame:
        path = os.path.join(folder, fname)
        if not os.path.isfile(path):
            return pd.DataFrame()
        try:
            return pd.read_csv(path)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    summary: Dict[str, Any] = {}
    summary_path = os.path.join(folder, "summary.json")
    if os.path.isfile(summary_path):
        try:
            with open(summary_path) as f:
                summary = json.load(f)
        except json.JSONDecodeError:
            summary = {}

    return {
        "folder": folder,
        "prior_date": prior_date,
        "score_movers": _safe_read("score_movers.csv"),
        "band_changes": _safe_read("band_changes.csv"),
        "quadrant_changes": _safe_read("quadrant_changes.csv"),
        "action_flag_changes": _safe_read("action_flag_changes.csv"),
        "new_funds": _safe_read("new_funds.csv"),
        "removed_funds": _safe_read("removed_funds.csv"),
        "summary": summary,
    }


def _load_model_overlay(
    runs_dir: str, run_date: str
) -> Optional[Dict[str, Any]]:
    """Return the model-holdings overlay artifacts for this run, if any."""
    folder = os.path.join(runs_dir, run_date, MODEL_OVERLAY_SUBDIR)
    if not os.path.isdir(folder):
        return None

    def _safe_read(fname: str) -> pd.DataFrame:
        path = os.path.join(folder, fname)
        if not os.path.isfile(path):
            return pd.DataFrame()
        try:
            return pd.read_csv(path)
        except pd.errors.EmptyDataError:
            return pd.DataFrame()

    scorecard_path = os.path.join(folder, "model_holdings_scorecard.csv")
    if not os.path.isfile(scorecard_path):
        return None

    meta: Dict[str, Any] = {}
    meta_path = os.path.join(folder, "overlay_metadata.json")
    if os.path.isfile(meta_path):
        try:
            with open(meta_path) as f:
                meta = json.load(f)
        except json.JSONDecodeError:
            meta = {}

    return {
        "folder": folder,
        "summary": _safe_read("model_summary.csv"),
        "scorecard": _safe_read("model_holdings_scorecard.csv"),
        "current_review": _safe_read("current_holdings_review.csv"),
        "research_candidates": _safe_read("research_candidates.csv"),
        "replacement_candidates": _safe_read("replacement_candidates.csv"),
        "metadata": meta,
    }


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_kv_block(
    ws: Worksheet,
    start_row: int,
    title: str,
    items: List[Tuple[str, Any]],
) -> int:
    """Write a labeled block of (key, value) rows. Returns next-row cursor."""
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    row = start_row + 1
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=_coerce_cell(value)).font = BODY_FONT
        row += 1
    return row + 1


def _write_readme(
    ws: Worksheet,
    run: Dict[str, Any],
    comparison: Optional[Dict[str, Any]],
    generated_at: str,
) -> None:
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 72

    meta = run.get("metadata", {}) or {}
    validation = run.get("validation", {}) or {}
    outputs = meta.get("outputs", {}) or {}
    dual = outputs.get("dual_score_table", {}) or {}
    inputs = meta.get("inputs", {}) or {}

    ws.cell(row=1, column=1, value="FundScore Audit Workbook").font = Font(
        name="Calibri", size=18, bold=True, color="FF1F4E78"
    )
    ws.cell(row=2, column=1, value=(
        "Excel 2019-compatible. Every number in this workbook is precomputed; "
        "no dynamic-array formulas, XLOOKUP, LET, FILTER, SORT, or UNIQUE are used."
    )).font = BODY_FONT

    cursor = 4
    cursor = _write_kv_block(ws, cursor, "Run Context", [
        ("Run date", meta.get("run_date") or run.get("run_date")),
        ("Generated at (UTC)", generated_at),
        ("Archive generated at", meta.get("generated_at")),
        ("Score system", meta.get("score_system")),
        ("Join how", meta.get("join_how")),
        ("Workbook schema version", WORKBOOK_SCHEMA_VERSION),
        ("Run folder", run.get("path")),
    ])
    cursor = _write_kv_block(ws, cursor, "Row Counts", [
        ("Scored funds", dual.get("row_count") or validation.get("row_count")),
        ("Joined (both systems)", validation.get("joined_count")),
        ("Missing 2023 score", validation.get("missing_score_2023")),
        ("Missing 2025 score", validation.get("missing_score_2025")),
    ])

    band_25 = validation.get("band_counts_2025", {}) or {}
    band_23 = validation.get("band_counts_2023", {}) or {}
    quad = validation.get("quadrant_counts", {}) or {}
    action = validation.get("action_flag_counts", {}) or {}
    cursor = _write_kv_block(ws, cursor, "Score Bands (2025)", [
        ("STRONG", band_25.get("STRONG", 0)),
        ("REVIEW", band_25.get("REVIEW", 0)),
        ("WEAK", band_25.get("WEAK", 0)),
    ])
    cursor = _write_kv_block(ws, cursor, "Score Bands (2023)", [
        ("STRONG", band_23.get("STRONG", 0)),
        ("REVIEW", band_23.get("REVIEW", 0)),
        ("WEAK", band_23.get("WEAK", 0)),
    ])
    cursor = _write_kv_block(ws, cursor, "Dual-Lens Quadrants", [
        ("Q1 Both Strong", quad.get("Q1_Both_Strong", 0)),
        ("Q2 Only 2025", quad.get("Q2_Only_2025", 0)),
        ("Q3 Only 2023", quad.get("Q3_Only_2023", 0)),
        ("Q4 Both Weak", quad.get("Q4_Both_Weak", 0)),
    ])
    cursor = _write_kv_block(ws, cursor, "Action Flags", [
        ("LEAD", action.get("LEAD", 0)),
        ("REVIEW", action.get("REVIEW", 0)),
        ("WATCH", action.get("WATCH", 0)),
        ("DROP", action.get("DROP", 0)),
    ])

    in_2025 = inputs.get("path_2025", {}) or {}
    in_2023 = inputs.get("path_2023", {}) or {}
    cursor = _write_kv_block(ws, cursor, "Input Files", [
        ("2025 path", in_2025.get("path")),
        ("2025 sha256", in_2025.get("sha256")),
        ("2025 size (bytes)", in_2025.get("size_bytes")),
        ("2023 path", in_2023.get("path")),
        ("2023 sha256", in_2023.get("sha256")),
        ("2023 size (bytes)", in_2023.get("size_bytes")),
    ])

    preflight = meta.get("preflight") or {}
    cursor = _write_kv_block(ws, cursor, "Intake Preflight", [
        ("Mode", preflight.get("mode")),
        ("Ran", preflight.get("ran")),
        ("Failed", preflight.get("failed")),
        ("Report path (relative)", preflight.get("report_relative_path")),
    ])

    if comparison is not None:
        summary = comparison.get("summary", {}) or {}
        cursor = _write_kv_block(ws, cursor, "Month-over-Month Comparison", [
            ("Prior run date", comparison.get("prior_date")),
            ("New funds", summary.get("new_fund_count")),
            ("Removed funds", summary.get("removed_fund_count")),
            ("Quadrant changes", summary.get("quadrant_change_count")),
            ("Action-flag changes", summary.get("action_flag_change_count")),
        ])

    cursor = _write_kv_block(ws, cursor, "Sheet Index", [
        (SHEET_METHODOLOGY, "Scoring methodology and band thresholds."),
        (SHEET_DATA_QUALITY, "Coverage, missing counts, validation summary."),
        (SHEET_SCORED, "Full canonical dual-score table (one row per fund)."),
        (SHEET_TOP50_2023, "Top 50 funds by 2023 Combined System score."),
        (SHEET_TOP50_2025, "Top 50 funds by 2025 Split System score."),
        (SHEET_TOP50_CONSENSUS, "Top 50 funds by Consensus_Rank."),
        (SHEET_DISAGREEMENT, "Funds where the 2023 and 2025 systems disagree."),
        (SHEET_WHAT_CHANGED, "Month-over-month headline (if comparison exists)."),
    ])


def _write_methodology(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 96
    ws.cell(row=1, column=1, value="Methodology").font = Font(
        name="Calibri", size=18, bold=True, color="FF1F4E78"
    )
    rows: List[Tuple[str, str]] = [
        ("Dual-system design", "The 2023 Combined System scores all funds on one rubric. The 2025 Split System separates Passive (10 metrics, cost + tracking) from Active (16 metrics, manager skill + risk-adjusted return)."),
        ("Within-category ranking", "All metrics are ranked within Morningstar category — Large Growth vs Large Growth peers — so category base rates do not dominate."),
        ("Pro-rating missing data", "Funds with partial history are scored on the metrics they have; the denominator adjusts so newer funds are not penalised."),
        ("Score bands", "STRONG: score >= 80.  REVIEW: 60 - 79.9.  WEAK: < 60."),
        ("Quadrants", "Q1 Both Strong: both >= 80.  Q2 Only 2025: 2025 >= 80, 2023 < 80.  Q3 Only 2023: 2023 >= 80, 2025 < 80.  Q4 Both Weak: everything else."),
        ("Action flags", "LEAD: both STRONG.  REVIEW: one STRONG.  WATCH: borderline / REVIEW / REVIEW.  DROP: both WEAK."),
        ("Consensus rank", "Dense rank on the average of Rank_2023 and Rank_2025 (1 = best)."),
        ("Validation targets", "OMCIX ~68.7 (Active). Workbook matches the canonical engine within +/- 0.1. Note: prior Passive validation target (SCHD ~69.6) was based on a now-removed 1.111 rescale that allowed Passive scores >100; both systems now produce 0-100 scores using the same available-weight normalization."),
        ("Excel 2019 compatibility", "No XLOOKUP / LET / FILTER / SORT / UNIQUE / SEQUENCE are used. All values are precomputed. Tables use the legacy ListObject format that ships with Excel 2007+."),
    ]
    for r, (label, value) in enumerate(rows, start=3):
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        cell = ws.cell(row=r, column=2, value=value)
        cell.font = BODY_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 36


def _write_data_quality(ws: Worksheet, validation: Dict[str, Any]) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.cell(row=1, column=1, value="Data Quality Summary").font = Font(
        name="Calibri", size=18, bold=True, color="FF1F4E78"
    )

    rows: List[Tuple[str, Any]] = [
        ("Row count", validation.get("row_count")),
        ("Joined count (both systems)", validation.get("joined_count")),
        ("Missing 2023 score", validation.get("missing_score_2023")),
        ("Missing 2025 score", validation.get("missing_score_2025")),
    ]
    for stat in ("score_2023", "score_2025", "score_gap"):
        block = validation.get(stat) or {}
        rows.append((f"{stat}.min", block.get("min")))
        rows.append((f"{stat}.max", block.get("max")))
        rows.append((f"{stat}.mean", block.get("mean")))
        rows.append((f"{stat}.missing", block.get("missing")))
    for cov in ("coverage_2023", "coverage_2025"):
        block = validation.get(cov) or {}
        rows.append((f"{cov}.mean", block.get("mean")))
        rows.append((f"{cov}.min", block.get("min")))
        rows.append((f"{cov}.missing", block.get("missing")))

    r_start = 3
    ws.cell(row=r_start, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=r_start, column=1).fill = HEADER_FILL
    ws.cell(row=r_start, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=r_start, column=2).fill = HEADER_FILL
    for i, (label, value) in enumerate(rows, start=r_start + 1):
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        ws.cell(row=i, column=2, value=_coerce_cell(value)).font = BODY_FONT

    cursor = r_start + len(rows) + 3

    def _write_count_block(title: str, counts: Dict[str, Any]) -> int:
        nonlocal cursor
        ws.cell(row=cursor, column=1, value=title).font = SECTION_FONT
        cursor += 1
        ws.cell(row=cursor, column=1, value="Bucket").font = HEADER_FONT
        ws.cell(row=cursor, column=1).fill = HEADER_FILL
        ws.cell(row=cursor, column=2, value="Count").font = HEADER_FONT
        ws.cell(row=cursor, column=2).fill = HEADER_FILL
        cursor += 1
        if not counts:
            ws.cell(row=cursor, column=1, value="(none)").font = BODY_FONT
            cursor += 1
        for key, val in sorted(counts.items()):
            ws.cell(row=cursor, column=1, value=str(key)).font = BODY_FONT
            ws.cell(row=cursor, column=2, value=_coerce_cell(val)).font = BODY_FONT
            cursor += 1
        cursor += 1
        return cursor

    _write_count_block("Score Bands 2023", validation.get("band_counts_2023") or {})
    _write_count_block("Score Bands 2025", validation.get("band_counts_2025") or {})
    _write_count_block("Quadrants", validation.get("quadrant_counts") or {})
    _write_count_block("Action Flags", validation.get("action_flag_counts") or {})
    _write_count_block("Fund Type", validation.get("fund_type_counts") or {})


def _write_what_changed(
    ws: Worksheet,
    comparison: Dict[str, Any],
    latest_date: str,
) -> None:
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 24
    ws.cell(row=1, column=1, value=f"What Changed — {latest_date} vs {comparison['prior_date']}").font = Font(
        name="Calibri", size=18, bold=True, color="FF1F4E78"
    )
    summary = comparison.get("summary", {}) or {}
    rows: List[Tuple[str, Any]] = [
        ("Latest run date", summary.get("latest_run_date") or latest_date),
        ("Prior run date", summary.get("prior_run_date") or comparison.get("prior_date")),
        ("Latest row count", summary.get("latest_row_count")),
        ("Prior row count", summary.get("prior_row_count")),
        ("New funds", summary.get("new_fund_count")),
        ("Removed funds", summary.get("removed_fund_count")),
        ("Quadrant changes", summary.get("quadrant_change_count")),
        ("Action-flag changes", summary.get("action_flag_change_count")),
    ]
    movers = summary.get("score_mover_count_by_metric") or {}
    for metric, count in sorted(movers.items()):
        rows.append((f"Score movers — {metric}", count))
    band = summary.get("band_change_count_by_column") or {}
    for col, count in sorted(band.items()):
        rows.append((f"Band changes — {col}", count))

    r = 3
    ws.cell(row=r, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = BODY_FONT
        ws.cell(row=r, column=2, value=_coerce_cell(value)).font = BODY_FONT
        r += 1

    ws.cell(row=r + 1, column=1, value="Detailed change tables").font = SECTION_FONT
    ws.cell(row=r + 2, column=1, value=(
        f"See the {SHEET_SCORE_MOVERS}, {SHEET_BAND_CHANGES}, "
        f"{SHEET_QUADRANT_CHANGES}, {SHEET_ACTION_CHANGES}, "
        f"{SHEET_NEW_FUNDS}, {SHEET_REMOVED_FUNDS} sheets."
    )).font = BODY_FONT


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build_audit_workbook(
    run: Dict[str, Any],
    out_path: str,
    runs_dir: Optional[str] = None,
    include_comparison: bool = True,
) -> str:
    """Write an Excel 2019-compatible audit workbook for a single run.

    Parameters
    ----------
    run : dict returned by ``run_archive.load_run`` or ``load_latest_run``.
    out_path : absolute or relative ``.xlsx`` path.
    runs_dir : root of the runs archive; used to discover any pre-computed
        ``run_comparison`` output that belongs to this run. If None we
        assume ``run["path"]`` is ``<runs_dir>/<run_date>`` and walk up one.
    include_comparison : if False, skip the comparison sheets even when
        the CSV bundle is present on disk.
    """
    table = run.get("table")
    if table is None:
        raise ValueError("run dict is missing a 'table' DataFrame")
    if not isinstance(table, pd.DataFrame):
        raise TypeError("run['table'] must be a pandas DataFrame")

    run_date = run.get("run_date") or (run.get("metadata") or {}).get("run_date") or "unknown"
    run_path = run.get("path")
    if runs_dir is None and run_path:
        runs_dir = os.path.dirname(os.path.abspath(run_path))

    comparison: Optional[Dict[str, Any]] = None
    if include_comparison and runs_dir and run_date != "unknown":
        comparison = _load_comparison_tables(runs_dir, run_date)

    overlay: Optional[Dict[str, Any]] = None
    if runs_dir and run_date != "unknown":
        overlay = _load_model_overlay(runs_dir, run_date)

    generated_at = dt.datetime.now(dt.timezone.utc).isoformat()

    wb = Workbook()
    wb.remove(wb.active)  # we create every sheet ourselves

    ws_readme = wb.create_sheet(SHEET_README)
    _write_readme(ws_readme, run, comparison, generated_at)

    ws_method = wb.create_sheet(SHEET_METHODOLOGY)
    _write_methodology(ws_method)

    ws_dq = wb.create_sheet(SHEET_DATA_QUALITY)
    _write_data_quality(ws_dq, run.get("validation") or {})

    ws_scored = wb.create_sheet(SHEET_SCORED)
    _add_table(ws_scored, table, "tbl_scored_funds")
    _apply_band_fills(ws_scored, table, ("Score_Band_2023", "Score_Band_2025"))

    top23 = _top_n_by_score(table, "Score_2023_Final")
    ws_t23 = wb.create_sheet(SHEET_TOP50_2023)
    _add_table(ws_t23, top23, "tbl_top50_2023")
    _apply_band_fills(ws_t23, top23, ("Score_Band_2023", "Score_Band_2025"))

    top25 = _top_n_by_score(table, "Score_2025_Final")
    ws_t25 = wb.create_sheet(SHEET_TOP50_2025)
    _add_table(ws_t25, top25, "tbl_top50_2025")
    _apply_band_fills(ws_t25, top25, ("Score_Band_2023", "Score_Band_2025"))

    topc = _top_n_by_consensus(table)
    ws_tc = wb.create_sheet(SHEET_TOP50_CONSENSUS)
    _add_table(ws_tc, topc, "tbl_top50_consensus")
    _apply_band_fills(ws_tc, topc, ("Score_Band_2023", "Score_Band_2025"))

    disagree = _disagreement_list(table)
    ws_d = wb.create_sheet(SHEET_DISAGREEMENT)
    _add_table(ws_d, disagree, "tbl_disagreement")
    _apply_band_fills(ws_d, disagree, ("Score_Band_2023", "Score_Band_2025"))

    if overlay is not None:
        summary_df = overlay.get("summary", pd.DataFrame())
        ws_ms = wb.create_sheet(SHEET_MODEL_SUMMARY)
        _add_table(ws_ms, summary_df, "tbl_model_summary")

        scorecard_df = overlay.get("scorecard", pd.DataFrame())
        ws_mh = wb.create_sheet(SHEET_MODEL_HOLDINGS)
        _add_table(ws_mh, scorecard_df, "tbl_model_holdings")
        _apply_band_fills(ws_mh, scorecard_df, ("Score_Band_2023", "Score_Band_2025"))

        review_df = overlay.get("current_review", pd.DataFrame())
        ws_mr = wb.create_sheet(SHEET_MODEL_REVIEW_LIST)
        _add_table(ws_mr, review_df, "tbl_model_review_list")
        _apply_band_fills(ws_mr, review_df, ("Score_Band_2023", "Score_Band_2025"))

        research_df = overlay.get("research_candidates", pd.DataFrame())
        ws_rc = wb.create_sheet(SHEET_RESEARCH_CANDIDATES)
        _add_table(ws_rc, research_df, "tbl_research_candidates")
        _apply_band_fills(ws_rc, research_df, ("Score_Band_2023", "Score_Band_2025"))

        replacement_df = overlay.get("replacement_candidates", pd.DataFrame())
        ws_rp = wb.create_sheet(SHEET_REPLACEMENT_CANDIDATES)
        _add_table(ws_rp, replacement_df, "tbl_replacement_candidates")

    if comparison is not None:
        ws_wc = wb.create_sheet(SHEET_WHAT_CHANGED)
        _write_what_changed(ws_wc, comparison, run_date)

        ws_sm = wb.create_sheet(SHEET_SCORE_MOVERS)
        _add_table(ws_sm, comparison.get("score_movers", pd.DataFrame()), "tbl_score_movers")

        ws_bc = wb.create_sheet(SHEET_BAND_CHANGES)
        _add_table(ws_bc, comparison.get("band_changes", pd.DataFrame()), "tbl_band_changes")

        ws_qc = wb.create_sheet(SHEET_QUADRANT_CHANGES)
        _add_table(ws_qc, comparison.get("quadrant_changes", pd.DataFrame()), "tbl_quadrant_changes")

        ws_ac = wb.create_sheet(SHEET_ACTION_CHANGES)
        _add_table(ws_ac, comparison.get("action_flag_changes", pd.DataFrame()), "tbl_action_changes")

        ws_nf = wb.create_sheet(SHEET_NEW_FUNDS)
        _add_table(ws_nf, comparison.get("new_funds", pd.DataFrame()), "tbl_new_funds")

        ws_rf = wb.create_sheet(SHEET_REMOVED_FUNDS)
        _add_table(ws_rf, comparison.get("removed_funds", pd.DataFrame()), "tbl_removed_funds")

    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".", exist_ok=True)
    wb.save(out_path)
    return out_path


def export_run(
    run_date: str,
    runs_dir: str = DEFAULT_RUNS_DIR,
    out_path: Optional[str] = None,
    include_comparison: bool = True,
) -> str:
    run = load_run(run_date, runs_dir=runs_dir)
    target = out_path or _default_workbook_path(run["path"])
    return build_audit_workbook(
        run=run,
        out_path=target,
        runs_dir=runs_dir,
        include_comparison=include_comparison,
    )


def export_latest_run(
    runs_dir: str = DEFAULT_RUNS_DIR,
    out_path: Optional[str] = None,
    include_comparison: bool = True,
) -> str:
    run = load_latest_run(runs_dir=runs_dir)
    target = out_path or _default_workbook_path(run["path"])
    return build_audit_workbook(
        run=run,
        out_path=target,
        runs_dir=runs_dir,
        include_comparison=include_comparison,
    )


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cmd_export(args: argparse.Namespace) -> int:
    if args.run_date:
        target = export_run(
            run_date=args.run_date,
            runs_dir=args.runs_dir,
            out_path=args.out,
            include_comparison=not args.no_comparison,
        )
    else:
        runs = list_runs(args.runs_dir)
        if not runs:
            print(f"error: no runs found under {args.runs_dir}")
            return 2
        target = export_latest_run(
            runs_dir=args.runs_dir,
            out_path=args.out,
            include_comparison=not args.no_comparison,
        )
    print(f"Wrote audit workbook: {target}")
    return 0


def _cli() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Generate an Excel 2019-compatible audit workbook from an "
            "archived scoring run."
        )
    )
    parser.add_argument(
        "--runs-dir", default=DEFAULT_RUNS_DIR,
        help="Directory that holds per-run folders (default: streamlit/runs).",
    )
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("export", help="Export the audit workbook for a run.")
    p.add_argument(
        "--run-date", default=None,
        help="YYYY-MM-DD of the run to export (default: latest archived run).",
    )
    p.add_argument(
        "--out", default=None,
        help=(
            "Output .xlsx path. Defaults to "
            "<runs-dir>/<run-date>/reports/fundscore_audit_workbook.xlsx."
        ),
    )
    p.add_argument(
        "--no-comparison", action="store_true",
        help="Skip the month-over-month comparison sheets even when present.",
    )
    p.set_defaults(func=_cmd_export)

    args = parser.parse_args()
    raise SystemExit(args.func(args))


if __name__ == "__main__":
    _cli()
